import helpers

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule

import matplotlib.pyplot as plt

from mpl_toolkits.mplot3d import Axes3D
from matplotlib import cm

import math

class Evaluator:
    def __init__(self, sim):
        self.sim = sim

    def collect_results(self, n_tries):
        n_customers = len(self.sim.customers)
        if n_customers == 2:
            return self.collect_results_2d(n_tries)
        elif n_customers == 3:
            return self.collect_results_3d(n_tries)
        else:
            raise NotImplementedError('Only 2d and 3d collect implemented!')

    def collect_results_3d(self, n_tries):
        print('running 3d simulation')
        scenarios = self.sim.generate_scenarios(n_tries)
        print(scenarios)
        results = []
        b1 = self.sim.C # int(math.floor(self.sim.C / self.sim.customers[0].consumptionPerReq) * self.sim.customers[0].consumptionPerReq)
        print('b1=', b1)
        for b2 in range(0, b1+1): #, int(self.sim.customers[1].consumptionPerReq)):
            for b3 in range(0, b2+1):  #, int(self.sim.customers[2].consumptionPerReq)):
                revenue = helpers.list_average(self.sim.run_simulation((b1,b2,b3), scenarios))
                res_tpl = (b2,b3,revenue)
                results.append(res_tpl)
        return results

    def compute_opt(self, result_tuples):
        opt_profit = 0
        opt_b = (0, 0)

        for tuple in result_tuples:
            if tuple[2] > opt_profit:
                opt_profit = tuple[2]
                opt_b = (tuple[0], tuple[1])

        for tuple in result_tuples:
            if tuple[2] == opt_profit and (tuple[0], tuple[1]) != opt_b:
                print('Alternative opt:\n', tuple)

        return (opt_b) + (opt_profit,)

    def collect_results_2d(self, n_tries):
        print('running 2d simulation')
        scenarios = self.sim.generate_scenarios(n_tries)
        return list(map(lambda b2: helpers.tuple_column_averages(self.sim.run_simulation(b2, scenarios)), range(self.sim.C+1)))

    def add_result_data(self, results, sheet, row_offset):
        ros = str(row_offset)
        sheet['A' + ros] = 'reslevel'
        for class_ix in range(1, self.sim.num_classes+1):
            sheet.cell(row=row_offset, column=1+class_ix).value = 'n_' + str(class_ix)
        sheet.cell(row=row_offset, column=1+self.sim.num_classes+1).value = 'profit'

        row = row_offset + 1
        ctr = 0
        for r in results:
            sheet.cell(row=row, column=1).value = ctr
            for class_ix in range(self.sim.num_classes):
                sheet.cell(row=row, column=1+class_ix+1).value = r[class_ix]
            sheet.cell(row=row, column=1+self.sim.num_classes+1).value = r[self.sim.num_classes]
            row += 1
            ctr += 1

        rule = ColorScaleRule(start_type='min', start_value=0, start_color='FFFF0000',
                              end_type='max', end_value=100, end_color='FF00FF00')
        sheet.conditional_formatting.add('D' + str(row_offset + 1) + ':D' + str(row), rule)

    def export_results_2d(self, results, filename):
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Results'
        sheet['A2'] = 'Optimal policy ='
        sheet['B2'] = self.sim.optimal_policy()
        self.add_result_data(results, sheet, 5)
        wb.save(filename)

    def export_results_3d(self, results, filename):
        print('serializing results 3d')
        helpers.tuple_list_to_csv_file(('b2', 'b3', 'revenue'), results, filename)

    def plot_results_3d(self, results):
        print('plotting results 3d')
        xs = helpers.extract_column(results, 0)
        ys = helpers.extract_column(results, 1)
        zs = helpers.extract_column(results, 2)
        fig = plt.figure()
        ax = fig.gca(projection='3d')
        ax.plot_trisurf(xs, ys, zs, cmap=cm.jet, linewidth=0.2)
        ax.set_xlabel('b2')
        ax.set_ylabel('b3')
        ax.set_zlabel('revenue')
        plt.show()

    def plot_results_2d(self, results, show_opt=False):
        print('plotting results 2d')
        def extract_col(col_ix):
            return list(map(lambda triple: triple[col_ix], results))

        def strint(f): return str(int(f))

        def add_info_text(ax2):
            sim = self.sim
            c_1 = sim.customers[0].consumptionPerReq
            c_2 = sim.customers[1].consumptionPerReq
            r_1 = sim.customers[0].revenuePerReq
            r_2 = sim.customers[1].revenuePerReq
            mu_1 = sim.customers[0].expD
            mu_2 = sim.customers[1].expD
            sigma_1 = sim.customers[0].devD
            sigma_2 = sim.customers[1].devD
            info_str = r'$r_1=' + strint(r_1) + ', r_2=' + strint(r_2) + ',$'
            info_str2 = r'$C=' + str(sim.C) + ', c_1=' + strint(c_1) + ', c_2=' + strint(c_2) + ',$'
            info_str3 = r'$\mu_1=' + str(mu_1) + ', \sigma_1=' + str(sigma_1) + ',$'
            info_str4 = r'$\mu_2=' + str(mu_2) + ', \sigma_2=' + str(sigma_2) + '$'
            ax2.text(2, 252, info_str)
            ax2.text(2, 242, info_str2)
            ax2.text(70, 252, info_str3)
            ax2.text(70, 242, info_str4)

        plt.rcParams.update({'font.size': 16})
        # plt.rc('text', usetex=True)

        xs = list(range(self.sim.C+1))

        fig, ax1 = plt.subplots()

        ax1.plot(xs, extract_col(0), 'r.', linewidth=2, label=r'$n_1(s_2)$')
        ax1.plot(xs, extract_col(1), 'g.', linewidth=2, label=r'$n_2(s_2)$')
        ax1.set_ylabel(r'$\#$ Auftr\"{a}ge angenommen')

        ax2 = ax1.twinx()
        ax2.plot(xs, extract_col(2), 'b.', linewidth=2, label=r'$u(s_2)$')
        ax2.set_ylabel(r'Erl\"{o}s in GE')

        add_info_text(ax2)

        if show_opt:
            ropt = self.sim.optimal_policy()
            plt.axvline(x=ropt, ymin=0, ymax=1, linewidth=2, color='gray')
            ax2.text(ropt - 2, 65, r'$s_2^*=' + str(ropt) + '$')

        ax1.legend(loc=(0.05, 0.3), handlelength=0, numpoints=1, frameon=False)
        ax2.legend(loc='upper right', handlelength=0, numpoints=1, frameon=False)
        ax1.set_xlabel(r'$s_2$')
        # plt.axes.text(right, bottom, 'center top', horizontalalignment='center', verticalalignment='top', transform=ax.transAxes)
        # plt.show()
        plt.savefig('test.pdf')

    def plot_littlewood_example(self):
        xs = list(range(self.sim.C+1))
        plt.axhline(self.sim.customers[1].revenuePerReq, 0, 1, label=r'$r_2='+str(int(self.sim.customers[1].revenuePerReq))+'$', linewidth=3, color='blue')
        ys = list(map(lambda x: self.sim.customers[0].revenuePerReq * (1 - helpers.dist_func(self.sim.customers[0].expD, self.sim.customers[0].devD, x)), xs))
        plt.rcParams.update({'font.size': 12})
        plt.rc('text', usetex=True)
        plt.xlabel(r'Restkapazit\"{a}t $x$')
        plt.ylabel(r'Erl\"{o}s')
        plt.plot(xs, ys, 'r.', markersize=12.0, label=r'Grenzerl\"{o}s $r_1 \cdot P(D_1 \geq x)$')
        s_opt = self.sim.optimal_policy()
        plt.axvline(s_opt, 0, 1, label=r'$s_2^*=' + str(s_opt) + '$', linewidth=3, color='gray')
        plt.legend(loc='upper right', handlelength=0, numpoints=1, frameon=True)
        #plt.show()
        plt.savefig('grenzerloes.pdf')